import os
from ciscoconfparse import CiscoConfParse
from pprint import pprint
import pandas as pd
import numpy as np
import re
import openpyxl
from openpyxl import load_workbook
import shutil

pd.options.display.width = None
pd.options.display.max_columns = None
pd.options.display.max_rows = None


####PART0 - Custom changes. Update before running the script

##INPUT -> Enter the Date report is generated on
reportdate="31Jul23"

##INPUT -> Enter name for the New sheet for this week in the format shown below
newsheet='24 Jul 23 to 31 Jul 23 (latest)'

##INPUT -> Download Raw Inventory from NP. Add Path for downloaded Unzipped RawInventory folder
basepath = '/Users/azile/Downloads/RawInventory_31Jul23'

##INPUT -> Dowload DAV Report from BDB. Enter path for the DAV Report - NP KEY: 179482 Group ID: 473451
## BDB link: https://scripts.cisco.com/ui/use/np_dav3
davpath = '/Users/azile/Downloads/179482-DAV-Jul_31_2023.xlsx'

##INPUT -> Enter path for Last week's report
lwpath='/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/IOS_IOSXE_ConfigDiffReport_24Jul23.xlsx'











dav = pd.ExcelFile(davpath)

####PART1 - Collects show command output for IOS/IOSXE devices####

def removecols(nstatus,naccess):
    if 'Mismatch' in nstatus and naccess =="FAIL":
        nstatus = "Collector Issue - SSH/Telnet Authentication Issue"
    return nstatus


def certcheck2(data):
    datalines = data.splitlines()
    if 'No changes were found' in data:
        ans = "MATCH"
    elif 'Unavailable' in data :
        ans = "Unavailable"
    elif "!The following order-dependent line(s) were re-ordered" in data:
        if "+" in data:
            ans="DOES NOT MATCH"
        else:
            ans = "MATCH"
    elif len(datalines) == 1:
        ans = "MATCH"
    elif len(datalines) > 1:
        if 'Contextual' in datalines[-1]:
            ans = "MATCH"
        else:
            if 'Uncompressed' in data:
                ans = "MATCH"
            else:
                if '+' in data:
                    ans = "DOES NOT MATCH"
                else:
                    ans = "DOES NOT MATCH"
    else:
        ans = "MATCH"
    return ans
def removefalsedevices(nswitch, ans):
    if ( nswitch.startswith('eorrci-pga1') or nswitch.startswith('eorrci-rces-') or nswitch.startswith('eorwdw-epc-')
      or nswitch.startswith('eorwdw-mgk') or nswitch.startswith('eorwdw-spt') or nswitch.startswith('eorwdw-stu') ) and ans == 'DOES NOT MATCH':
        newans = 'MATCH'
    else:
        newans = ans
    return newans
def removecrypto(nswitch,data):
    if len(data)>0:
        datalines = data.splitlines()
    else:
        datalines=data

    parse = CiscoConfParse(datalines)

    for obj in parse.find_objects(r"^nvram.*cer"):
        obj.delete_children_matching(r'.*')
    parse.commit()
    parse.delete_lines(linespec="^.*nvram.*cer")
    parse.commit()

    for obj in parse.find_objects(r"^.*crypto pki"):
        obj.delete_children_matching(r'.*')
    parse.commit()
    parse.delete_lines(linespec="^.*crypto pki")
    parse.commit()

    for obj in parse.find_objects(r"^.*crypto ca"):
        obj.delete_children_matching(r'.*')
    parse.commit()
    parse.delete_lines(linespec="^.*crypto ca")
    parse.commit()

    withoutcrypto = []
    for line in parse.ioscfg:
         withoutcrypto.append(line)
    withoutcryptonew = '\n'.join(withoutcrypto)

    return withoutcryptonew
def removentp(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)
    for obj in parse.find_objects(r'.*ntp clock.*'):
        obj.delete_children_matching(r'.*')
    parse.commit()
    parse.delete_lines(linespec="ntp clock")

    withoutntp = []
    for line in parse.ioscfg:
        withoutntp.append(line)
    parse.commit()

    withoutntpnew = '\n'.join(withoutntp)
    return withoutntpnew
def removesticky(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)

    crypto_maps = parse.find_objects(r"^interface")
    for c_map in crypto_maps:
        for child in c_map.children:
            childtxt=child.text
            if 'sticky' in childtxt:
                child.delete()
    parse.commit()
    nochild=parse.find_parents_wo_child(r'interface.*', '.*')

    for obj in parse.find_objects(r'interface.*'):
        if obj.text in nochild:
            obj.delete() # removing parents with no children here
    parse.commit()

    withoutsticky = []
    for line in parse.ioscfg:
        withoutsticky.append(line)
    parse.commit()

    withoutsickynew = '\n'.join(withoutsticky)
    return withoutsickynew
def remove3exceptions(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)

    #remove snmp-server engineID.*
    for obj in parse.find_objects(r'snmp-server engineID.*'):
        #print(obj.text)
        obj.delete()
    parse.commit()

    #remove radius-server host.*
    for obj in parse.find_objects(r'radius-server host.*'):
        #print(obj.text)
        obj.delete()
    parse.commit()

    #remove enable secret.*
    for obj in parse.find_objects(r'enable secret.*'):
        #print(obj.text)
        obj.delete()
    parse.commit()

    wo3exceptions = []
    for line in parse.ioscfg:
        wo3exceptions.append(line)
    parse.commit()

    wo3exceptions_new = '\n'.join(wo3exceptions)
    return wo3exceptions_new
def removebanner(data):
    datalines = data.splitlines()
    #pprint(datalines)
    parse = CiscoConfParse(datalines)

    bannerdata = ['=============================================================','Computer System Terms of Use',
 'This system is only for authorized use.  Using this system',
 'means all of your activity and communications on it,',
 'including electronic mail and Internet use, may be',
 'monitored, recorded and disclosed subject to applicable law',
 'and the Company computer usage and security policy. The use',
 'of passwords does not constitute any promise of',
 'confidentiality regarding any such activity or communication',
 'created, accessed or stored on this system.  By using this',
 'system, you are expressly consenting to these Terms of Use.',
 'Any unauthorized or inappropriate use of this system may',
 'subject the user to disciplinary action up to and including',
 'termination.',
 '=============================================================',
 '^C',
 '====================================',
 'Computer System Terms of Use',
 'This system is only for authorized use. Using this system',
 'means all of your activity and communications on it,',
 'including electronic mail and Internet use, may be monitored,',
 'recorded and disclosed subject to applicable law and the',
 'Company computer usage and security policy. The use of',
 'passwords does not constitute any promise of confidentiality',
 'regarding any such activity or communication created,',
 'accessed or stored on this system. By using this system, you',
 'are expressly consenting to these Terms of Use. Any',
 'unauthorized or inappropriate use of this system may subject',
 'the user to disciplinary action up to and including',
 'termination.',
 '====================================',
 '^C',
 'This system is only for authorized use.  Using this system   ',
 'means all of your activity and communications on it, ',
 'including electronic mail and Internet use, may be ',
 'monitored, recorded and disclosed subject to applicable law ',
 'and the Company computer usage and security policy. The use ',
 'of passwords does not constitute any promise of ',
 'confidentiality regarding any such activity or communication ',
 'created, accessed or stored on this system.  By using this',
 'system, you are expressly consenting to these Terms of Use. ',
 'Any unauthorized or inappropriate use of this system may ',
 'subject the user to disciplinary action up to and including ',
 'termination.',
 '============================================================',
 'of passwords does not constitute any promise of ',
 'subject the user to disciplinary action up to and including ',
                  '============================================================= ',
                  'Computer System Terms of Use ',
                  'This system is only for authorized use. Using this system ',
                  'created, accessed or stored on this system. By using this ',
                  'termination. ',
                  '============================================================= ',
                  "% Invalid input detected at '^' marker.",
                                                 '       ^',
                  'system, you are expressly consenting to these Terms of Use',
                  'termination.^C',
                  '===========================================',
                  'This system is only for authorized use. Using',
                  'this system means all of your activity and',
                  'communications on it, including electronic',
                  'mail and Internet use, may be monitored,',
                  'recorded and disclosed subject to applicable',
                  'law and the Company computer usage and',
                  'security policy. The use of passwords does not',
                  'constitute any promise of confidentiality',
                  'regarding any such activity or communication',
                  'created, accessed or stored on this system. By',
                  'using this system, you are expressly',
                  'consenting to these Terms of Use. Any',
                  'unauthorized or inappropriate use of this',
                  'system may subject the user to disciplinary',
                  'action up to and including termination.'
                  ]

    #remove banner
    for obj in parse.find_objects(r'.*banner.*'):
        #print(obj.text)
        obj.delete_children_matching(r'.*')
    parse.commit()
    parse.delete_lines(linespec="banner")

    #remove bannerdata
    for obj in parse.find_objects(r'.*'):
        line=obj.text
        #line2=line.trim()
        if line in bannerdata:
            obj.delete()
    parse.commit()

    wobanner = []
    for line in parse.ioscfg:
        wobanner.append(line)
    parse.commit()
    wobanner_new = '\n'.join(wobanner)

    return wobanner_new
def noswitchport(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)

    for intf_obj in parse.find_objects_w_child('^interface', '^ -switchport$'):
        intf_obj.delete()
    parse.commit()

    crypto_maps = parse.find_objects(r"^interface")
    for c_map in crypto_maps:
        for child in c_map.children:
            childtxt=child.text
            # print(child.text)
            if childtxt==" +switchport":
                child.delete()
    parse.commit()

    withoutcrypto = []
    for line in parse.ioscfg:
        withoutcrypto.append(line)
    parse.commit()
    withoutcryptonew = '\n'.join(withoutcrypto)

    return withoutcryptonew
def noaaaserverkey(data):
    datalines = data.splitlines()
    return data


path=os.path.join(basepath, 'CLI')
dir_list1 = os.listdir(path)

newlist=[]
for item in dir_list1:
    if item != '.DS_Store':
        newlist.append(item)

finallist1 = []
#arraylist=["enaaot-dur-mc401-sw", "eeutds-fr852m-gw", "wbucrp-234bv1-sw","wnadvc-d-e1115-0-sw","aandlr-tda1-wc","eorwdw-cor-fd-sw"]

print("PART 1 - Started...",len(newlist)," ")

cgk=0
for i in newlist:
    dict = {}
    fpath = os.path.join(path, i)  # fpath for devices
    dict['switch'] = i
    dev_list2 = os.listdir(fpath)
    os.system('find . -name ".DS_Store" -delete')
    dev_list2 = os.listdir(fpath)

    for j in dev_list2:
        dpath = os.path.join(fpath, j)  # dpath for txt file
        print(cgk, " ", dpath)
        cgk=cgk+1
        with open(dpath) as f:
            sconf = f.read()
            if i != "eorwdw-td1n3-1-sw": #eeucrp-hubuvoice02-vg
                dict['detailed_diffs'] = sconf

    finallist1.append(dict)
df1 = pd.DataFrame(finallist1)

df1.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/df1_'+reportdate+'_temp.csv', index=False)
#df1=pd.read_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/df1_'+reportdate+'_temp.csv')

pprint(df1.head())
print(df1.shape,"DF1")
df1 = df1.dropna()

df1['no crypto config'] = df1['detailed_diffs']

df1['no crypto config'] = np.vectorize(removecrypto)(df1['switch'], df1['detailed_diffs']) #removes crypto
df1['no ntp'] = df1['no crypto config'].apply(removentp) #hides ntp sections
df1['no sticky'] = df1['no ntp'].apply(removesticky) #hides sticky mac addresses from int
df1['no 3exceptions'] = df1['no sticky'].apply(remove3exceptions) #removes snmp-server engineId, remove radius-server host, # enable secret
df1['no banner'] = df1['no 3exceptions'].apply(removebanner) #removes banner data
df1['no switchport']=df1['no banner'].apply(noswitchport) #removes switchport
df1['no aaa dynamic author']=df1['no switchport'].apply(noaaaserverkey)

#validates matching or not
df1['startup_matches_running'] = df1['no aaa dynamic author'].apply(certcheck2)
dfCopy = df1.copy()

#rename columns
dfCopy = dfCopy.drop('detailed_diffs', 1) # remove detailed diffs (contains crypto)
dfCopy = dfCopy.drop('no ntp', 1)
dfCopy = dfCopy.drop('no crypto config', 1)
dfCopy = dfCopy.drop('no sticky', 1)
dfCopy = dfCopy.drop('no 3exceptions', 1)
dfCopy = dfCopy.drop('no banner', 1)
dfCopy = dfCopy.drop('no switchport', 1)
dfCopy.rename(columns={'no aaa dynamic author': 'detailed_diffs'}, inplace=True)

print("#################Part 1 - Data ", len(dfCopy), "################")
print(dfCopy.shape)
pprint(dfCopy.head())

#dfCopy.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/df1_part1_'+reportdate+'_tempv2.csv', index=False)






####PART2 - Collects timestamp and user data ####
def lastconfig(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)
    lines=[]
    for obj in parse.find_objects(r'.*Last configuration change.*'):
        lconfig = obj.text
        x = re.search(r"Last configuration change at (.*)", lconfig)
        lconfig = x.group(1)
        lines.append(lconfig)

    linesnew = '\n'.join(lines)
    return linesnew

def nvramconfig(data):
    datalines = data.splitlines()
    parse = CiscoConfParse(datalines)

    lines = []
    for obj in parse.find_objects(r'.*NVRAM config last updated.*'):
        nconfig = obj.text
        x = re.search(r"NVRAM config last updated at (.*)", nconfig)
        nconfig = x.group(1)
        lines.append(nconfig)

    yo=[]
    if len(lines)>1:
        yo.append(lines[0])
    else:
        yo = lines

    linesnew = '\n'.join(yo)
    return linesnew

def collect_user(data):
    datalines = data.split(' ')
    if 'by' in datalines:
        lastuser = datalines[-1]
    else:
        lastuser = ""
    return lastuser

def keepdates(data):
    datalines = data.split(' ')

    if 'by' in datalines:
        date = datalines[:-2]
    else:
        date = datalines

    datenew=' '.join(date)
    return datenew


path2=os.path.join(basepath, 'Config')
dir_list2 = os.listdir(path2)
newlist2=[]

for item in dir_list2:
    if item != '.DS_Store':
        newlist2.append(item)

finallist2 = []
print(len(newlist2),"PART 2 - Started...")

cgk2=0
for i in newlist2:
    dict2 = {}
    fpath2 = os.path.join(path2, i)  # fpath for devices

    dev_list22 = os.listdir(fpath2)
    dict2['switch'] = i

    os.system('find . -name ".DS_Store" -delete')
    for j in dev_list22:
        dpath2 = os.path.join(fpath2, j)  # dpath for txt file run/start
        print("running", cgk2, " ", dpath2)
        with open(dpath2) as f:
            sconf2 = f.read()[0:5000]
            if 'startup' in dpath2:
                dict2['startup_LastConfigChange'] = sconf2
                dict2['User1'] = sconf2
                dict2['startup_LastNVRAMConfigChange'] = sconf2
                dict2['User2'] = sconf2

            if 'running' in dpath2:
                dict2['running_LastConfigChange'] = sconf2
                dict2['User3'] = sconf2
                dict2['running_LastNVRAMConfigChange'] = sconf2
                dict2['User4'] = sconf2
        cgk2=cgk2+1
    finallist2.append(dict2)

df2 = pd.DataFrame(finallist2)
df2 = df2.fillna("Not Available")
pprint(df2.head())

df2.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/df2_'+reportdate+'_temp.csv', index=False)

print("Reading df...")
#df2=pd.read_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/df2_'+reportdate+'_temp.csv')

#collect LastConfigChange
df2['startup_LastConfigChange']=df2['startup_LastConfigChange'].apply(lastconfig)
df2['User1']=df2['startup_LastConfigChange'].apply(collect_user)
df2['running_LastConfigChange']=df2['running_LastConfigChange'].apply(lastconfig)
df2['User3']=df2['running_LastConfigChange'].apply(collect_user)

#collect nvramChange
df2['startup_LastNVRAMConfigChange']=df2['startup_LastNVRAMConfigChange'].apply(nvramconfig)
df2['User2']=df2['startup_LastNVRAMConfigChange'].apply(collect_user)
df2['running_LastNVRAMConfigChange']=df2['running_LastNVRAMConfigChange'].apply(nvramconfig)
df2['User4']=df2['running_LastNVRAMConfigChange'].apply(collect_user)

#keep only dates
df2['startup_LastConfigChange']=df2['startup_LastConfigChange'].apply(keepdates)
df2['running_LastConfigChange']=df2['running_LastConfigChange'].apply(keepdates)
df2['startup_LastNVRAMConfigChange']=df2['startup_LastNVRAMConfigChange'].apply(keepdates)
df2['running_LastNVRAMConfigChange']=df2['running_LastNVRAMConfigChange'].apply(keepdates)


df2 = df2.replace(r'^\s*$', np.nan, regex=True)
df2=df2.fillna('N/A')
joindf = pd.merge(dfCopy, df2,  how='left', left_on=['switch'], right_on = ['switch'])

#joindf.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/thisweek_merged_'+reportdate+'.csv', index=False)
print("#################Part 2 - Data ", len(joindf), "################")
print(joindf.head())









####PART3 - Compares with last week's report####

# Combining last week with this week
# INPUT -> Enter path for Last week's report
#lwpath='/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/IOS_IOSXE_ConfigDiffReport_8May23.xlsx'
import openpyxl
from openpyxl import load_workbook

length=len(lwpath)-5
newlwpath=lwpath[0:length]+"_copy.xlsx"
shutil.copy(lwpath, newlwpath)

finalpath = newlwpath
print(finalpath)
book = load_workbook(finalpath)

allsheets=book.sheetnames

cleanedall=[]
for i in range (len(allsheets)):
    str_sheet_name = str(allsheets[i]) # converting to str
    #print(str_sheet_name)
    if "Sheet" not in str_sheet_name:
        #print(str_sheet_name)
        cleanedall.append(str_sheet_name)

deletedsheet=cleanedall[2]
lastweek_sheet=cleanedall[-1]

print("%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%")
print(lastweek_sheet)
print("%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%")

xls_full = pd.ExcelFile(lwpath) #update name

# INPUT -> Enter name of the latest sheet in last week's report
df_lastweek = pd.read_excel(xls_full, lastweek_sheet)

# df_lastweek = pd.read_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/ios_iosxedata_14Nov2022_lastweek.csv', encoding='UTF-8')
df_lastweek = df_lastweek.drop(['LastWeek_Diffs', 'LastWeek_MatchCheck','Status','running_LastConfigChange','User3','running_LastNVRAMConfigChange','User4','startup_LastConfigChange','User1','startup_LastNVRAMConfigChange','User2'], axis=1) #drop irrelevant columns

# df_lastweek.rename(columns={'LastWeek_Diffs': 'detailed_diffs_lastweek','ThisWeek_Diffsd': 'etailed_diffs_thisweek','ThisWeek_MatchCheck': 'startup_matches_running_thisweek'}, inplace=True)
df_lastweek.rename(columns={'LastWeek_Diffs': 'detailed_diffs','ThisWeek_Diffs': 'detailed_diffs','ThisWeek_MatchCheck': 'startup_matches_running','device':'switch'}, inplace=True)

# print("\nLast week 10: ", len(df_lastweek))
# print(df_lastweek.head())
#
# print("\nThis week 10: ", len(joindf))
# print(joindf.head())

combined_df = pd.merge(joindf, df_lastweek,  how='outer', left_on=['switch'], right_on = ['switch'],suffixes=('_thisweek', '_lastweek'))
combined_df[['startup_matches_running_lastweek', 'detailed_diffs_lastweek']] = combined_df[['startup_matches_running_lastweek','detailed_diffs_lastweek']].fillna("N/A this week")
combined_df[['startup_matches_running_thisweek', 'detailed_diffs_thisweek']] = combined_df[['startup_matches_running_thisweek','detailed_diffs_thisweek']].fillna("N/A this week")

combined_df['Status'] =                         combined_df.apply(lambda x: "Consistently Match" if (x['startup_matches_running_lastweek'] == x['startup_matches_running_thisweek'] == "MATCH")
                                                else ("Newly added to the report. Match." if ( x['startup_matches_running_lastweek']=="N/A this week" and x['startup_matches_running_thisweek'] == "MATCH")
                                                else ("Newly added to the report. Mismatch. " if (x['startup_matches_running_lastweek']=="N/A this week" and x['startup_matches_running_thisweek'] == "DOES NOT MATCH")
                                                else ("Device Unreachable this week" if (x['startup_matches_running_lastweek']!="N/A this week" and x['startup_matches_running_thisweek'] == "N/A this week") #newcheck
                                                else ("Saved since last week. Match." if  (x['startup_matches_running_lastweek'] == "DOES NOT MATCH" and x['startup_matches_running_thisweek'] == "MATCH")
                                                else ("Unsaved since last week. Mismatch." if (x['startup_matches_running_lastweek'] == "MATCH" and x['startup_matches_running_thisweek'] == "DOES NOT MATCH")
                                                else ("Consistently Mismatch" if (x['startup_matches_running_lastweek'] == "DOES NOT MATCH" and x['startup_matches_running_thisweek'] == "DOES NOT MATCH")
                                                else ("Newly added to the report. Match." if ( x['startup_matches_running_lastweek']=="N/A this week" and x['startup_matches_running_thisweek'] == "MATCH")
                                                else "Verify" )  ) ) )))),
                                            axis=1)

combined_df.rename(columns={'detailed_diffs_lastweek': 'LastWeek_Diffs','detailed_diffs_thisweek': 'ThisWeek_Diffs','startup_matches_running_lastweek': 'LastWeek_MatchCheck','startup_matches_running_thisweek': 'ThisWeek_MatchCheck','switch':'device'}, inplace=True)
'''
match match                         ok
n/a   match                         OK.new device
n/a   dn_match                      Please Save configs (new device)

match/dn_match n/a                  Device Unreachable this week

dn_match match                      ok (Saved this week)
match dn_match                      Please Save configs (Updated this week)
dn_match dn_match                   Please Save configs (Repeat)
n/a n/a                             Delete this
'''

combined_df = combined_df[['device', 'LastWeek_Diffs', 'LastWeek_MatchCheck', 'ThisWeek_Diffs', 'ThisWeek_MatchCheck','Status', 'running_LastConfigChange', 'User3', 'running_LastNVRAMConfigChange','User4','startup_LastConfigChange','User1','startup_LastNVRAMConfigChange','User2']]
combined_df=combined_df.fillna('N/A')
combined_df = combined_df[combined_df.Status != 'Verify']
print("#################Part 3 - Combined Data ", len(combined_df), "################")
print(combined_df.head())
print("------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------")

combined_df.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/part3_comparison_'+reportdate+'.csv', index=False)







####PART4 - Combines with DAV Report####

# INPUT -> Enter path for the DAV Report
#dav = pd.ExcelFile('/Users/azile/Downloads/179482-DAV-May_15_2023.xlsx')
dav_thisweek = pd.read_excel(dav, '473451')
dav_thisweek = dav_thisweek[['deviceName', 'Access status','configTime']]
dav_thisweek.rename(columns={'deviceName':'device'}, inplace=True)


#combined_df=pd.read_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/final_combineddf_'+reportdate+'.csv')
combined_df_dav = pd.merge(combined_df, dav_thisweek, how="left", on="device")
combined_df_dav['Status'] = np.vectorize(removecols)(combined_df_dav['Status'], combined_df_dav['Access status'])
combined_df_dav.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/part4_'+reportdate+'.csv', index=False)
combined_df_dav = combined_df_dav.drop('Access status', axis=1)
combined_df_dav = combined_df_dav.drop('configTime', axis=1)
print("#################Part 4 - Combined Data with DAV ", len(combined_df_dav), "################")
print(combined_df_dav.head())

#combined_df_dav.to_csv('/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/DAV_final_combineddf__'+reportdate+'.csv', index=False)







####PART5 - Creates final Excel report####
#################
##INPUT -> Enter name of the first sheet in last week's report. It will be removed/deleted.
#deletedsheet='24 Apr 23 to 1 May 23'

##INPUT -> Enter name of the last sheet in last week's report
#lastweek_sheet='15 May 23 to 23 May 23 (latest)'

#################
writer = pd.ExcelWriter(finalpath, engine = 'openpyxl')
writer.book = book


# INPUT -> Enter Name for the new sheet for this week
combined_df_dav.to_excel(writer, sheet_name = newsheet ,index = False)
writer.close()

# INPUT -> Update the last weeks sheet's name
ss = openpyxl.load_workbook(finalpath)
ss_sheet = ss.get_sheet_by_name(lastweek_sheet)
lastweek_sheet_new=lastweek_sheet[:-9]
ss_sheet.title = lastweek_sheet_new

# INPUT -> Delete the sheet for the first week. Enter name of sheet here.
pfd = ss[deletedsheet]
ss.remove(pfd)

#highlight column status


#rename new excel/copy
ss.save(r'/Users/azile/Library/CloudStorage/OneDrive-Cisco/DISNEY/CONFIG-DIFF/IOS_IOSXE_ConfigDiffReport_'+reportdate+'.xlsx')
print('IOS_IOSXE_ConfigDiffReport_'+reportdate+'.xlsx')
print("--------------------------------_END -----------------------")
print("--------------------------------_END -----------------------")


